#!/usr/bin/env python3
"""
Send Complete Reports via Email
Sends call productivity report, fax analysis, and comprehensive summary
"""

import os
import sys
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def load_env():
    """Load environment variables from .env file"""
    env_file = '.env'
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()

def get_fax_summary(date_str):
    """Get fax summary from the report"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f'exports/fax_analysis_{date_str}.xlsx')
        ws = wb['Faxes by Sender']
        
        fax_senders = []
        total_sent = 0
        total_received = 0
        
        for row in list(ws.iter_rows())[1:]:  # Skip header
            name = row[0].value
            ext = row[1].value
            sent = row[2].value or 0
            received = row[3].value or 0
            
            total_sent += sent
            total_received += received
            
            if ext and sent > 0:  # Only internal users who sent faxes
                fax_senders.append({
                    'name': name,
                    'ext': ext,
                    'sent': sent
                })
        
        fax_senders.sort(key=lambda x: x['sent'], reverse=True)
        
        return {
            'senders': fax_senders,
            'total_sent': total_sent,
            'total_received': total_received,
            'total': total_sent + total_received
        }
    except Exception as e:
        print(f"⚠️  Could not load fax summary: {e}")
        return None

def get_call_summary(date_str):
    """Get call summary from the report"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f'exports/{date_str}-nightangle-calls.xlsx')
        ws = wb.active
        
        total_calls_made = 0
        total_calls_received = 0
        total_minutes = 0
        all_callers = []
        
        for row in list(ws.iter_rows())[1:]:  # Skip header
            employee = row[0].value
            ext = row[1].value
            calls_received = row[5].value or 0
            calls_made = row[6].value or 0
            total_calls = row[7].value or 0
            minutes = row[12].value or 0
            
            if employee and ext:
                total_calls_made += calls_made
                total_calls_received += calls_received
                total_minutes += minutes
                
                if total_calls > 0:
                    all_callers.append({
                        'name': employee,
                        'ext': ext,
                        'calls_made': calls_made,
                        'calls_received': calls_received,
                        'calls': total_calls,
                        'minutes': minutes
                    })
        
        # Sort by total calls
        all_callers.sort(key=lambda x: x['calls'], reverse=True)
        
        return {
            'total_made': total_calls_made,
            'total_received': total_calls_received,
            'total_calls': total_calls_made + total_calls_received,
            'total_minutes': total_minutes,
            'all_callers': all_callers  # Return ALL callers
        }
    except Exception as e:
        print(f"⚠️  Could not load call summary: {e}")
        return None

def create_email_body(date_str, call_summary, fax_summary):
    """Create HTML email body with comprehensive analysis"""
    
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
            h2 {{ color: #34495e; margin-top: 25px; border-bottom: 2px solid #95a5a6; padding-bottom: 5px; }}
            h3 {{ color: #7f8c8d; margin-top: 20px; }}
            table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
            th {{ background-color: #3498db; color: white; padding: 12px; text-align: left; }}
            td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
            tr:hover {{ background-color: #f5f5f5; }}
            .summary-box {{ background-color: #ecf0f1; padding: 15px; border-radius: 5px; margin: 15px 0; }}
            .highlight {{ background-color: #fff3cd; padding: 2px 5px; border-radius: 3px; }}
            .metric {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
            .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #95a5a6; color: #7f8c8d; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>📊 RingCentral Activity Report - {date_str}</h1>
        
        <div class="summary-box">
            <h2>📈 Executive Summary</h2>
            <p>This report provides a comprehensive analysis of all call and fax activity for <strong>{date_str}</strong>.</p>
        </div>
    """
    
    # Call Summary
    if call_summary:
        html += f"""
        <h2>📞 Voice Call Activity</h2>
        <div class="summary-box">
            <table style="width: auto; border: none;">
                <tr style="border: none;">
                    <td style="border: none;"><strong>Total Calls Made:</strong></td>
                    <td style="border: none;"><span class="metric">{call_summary['total_made']:,}</span></td>
                    <td style="border: none; padding-left: 30px;"><strong>Total Calls Received:</strong></td>
                    <td style="border: none;"><span class="metric">{call_summary['total_received']:,}</span></td>
                </tr>
                <tr style="border: none;">
                    <td style="border: none;"><strong>Total Calls:</strong></td>
                    <td style="border: none;"><span class="metric">{call_summary['total_calls']:,}</span></td>
                </tr>
            </table>
        </div>
        
        <h3>📞 All Employees - Call Activity</h3>
        <table>
            <tr>
                <th>Employee Name</th>
                <th>Extension</th>
                <th>Calls Made<br>(Outbound)</th>
                <th>Calls Received<br>(Inbound)</th>
                <th>Total Calls</th>
            </tr>
        """
        
        for caller in call_summary['all_callers']:
            html += f"""
            <tr>
                <td>{caller['name']}</td>
                <td>{caller['ext']}</td>
                <td>{caller['calls_made']}</td>
                <td>{caller['calls_received']}</td>
                <td><strong>{caller['calls']}</strong></td>
            </tr>
            """
        
        html += "</table>"
    
    # Fax Summary
    if fax_summary:
        html += f"""
        <h2>📠 Fax Activity</h2>
        <div class="summary-box">
            <table style="width: auto; border: none;">
                <tr style="border: none;">
                    <td style="border: none;"><strong>Total Faxes Sent (Successful):</strong></td>
                    <td style="border: none;"><span class="metric">{fax_summary['total_sent']}</span></td>
                    <td style="border: none; padding-left: 30px;"><strong>Total Faxes Received (Successful):</strong></td>
                    <td style="border: none;"><span class="metric">{fax_summary['total_received']}</span></td>
                </tr>
                <tr style="border: none;">
                    <td style="border: none;"><strong>Total Faxes:</strong></td>
                    <td style="border: none;"><span class="metric">{fax_summary['total']}</span></td>
                    <td style="border: none; padding-left: 30px;"><strong>Active Fax Users:</strong></td>
                    <td style="border: none;"><span class="metric">{len(fax_summary['senders'])}</span></td>
                </tr>
            </table>
            <p style="margin-top: 10px; font-size: 12px; color: #666;">
                <em>Note: Only successful faxes are counted. Unsuccessful attempts (Busy, No Answer, Failed) are excluded.</em>
            </p>
        </div>
        
        <h3>📤 Employees Who Sent Faxes (Successful Only)</h3>
        <table>
            <tr>
                <th>Rank</th>
                <th>Employee Name</th>
                <th>Extension</th>
                <th>Faxes Sent</th>
            </tr>
        """
        
        for i, sender in enumerate(fax_summary['senders'], 1):
            highlight = ' class="highlight"' if i <= 3 else ''
            html += f"""
            <tr{highlight}>
                <td>{i}</td>
                <td><strong>{sender['name']}</strong></td>
                <td>{sender['ext']}</td>
                <td><strong>{sender['sent']}</strong></td>
            </tr>
            """
        
        html += "</table>"
    
    # Footer
    html += f"""
        <div class="footer">
            <h3>📋 Attached Reports</h3>
            <ul>
                <li><strong>{date_str}-nightangle-calls.xlsx</strong> - Complete call productivity report with all metrics</li>
                <li><strong>fax_analysis_{date_str}.xlsx</strong> - Detailed fax analysis with sender information and timestamps</li>
            </ul>
            
            <p><strong>Note:</strong> The fax analysis report contains two sheets:</p>
            <ul>
                <li>Sheet 1: Summary by sender</li>
                <li>Sheet 2: Complete fax log with timestamps and details</li>
            </ul>
            
            <p style="margin-top: 20px;">
                <em>Report generated on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')}</em><br>
                <em>RingCentral Analytics System</em>
            </p>
        </div>
    </body>
    </html>
    """
    
    return html

def send_email_with_reports(date_str):
    """Send email with both reports and comprehensive analysis"""
    print("📧 Preparing to send comprehensive report email...")
    
    load_env()
    
    # Email configuration
    sender_email = os.getenv('SENDER_EMAIL', 'nvenu@solifetec.com')
    
    # Get recipient emails from environment variable (comma-separated)
    # Example: RECIPIENT_EMAILS=email1@domain.com,email2@domain.com,email3@domain.com
    recipient_emails_str = os.getenv('RECIPIENT_EMAILS', '')
    
    if recipient_emails_str:
        # Use emails from .env file
        receiver_emails = [email.strip() for email in recipient_emails_str.split(',')]
    else:
        # Fallback to default emails if not configured
        receiver_emails = [
            # "dogden@HomeCareForYou.com",  # Commented out - configure in .env file
            # "DrBrar@HomeCareForYou.com",  # Commented out - configure in .env file
            "nvenu@solifetec.com"
        ]
        print("⚠️  Warning: RECIPIENT_EMAILS not set in .env file, using default")
    
    password = os.getenv('EMAIL_PASSWORD')
    
    if not password:
        print("❌ Error: EMAIL_PASSWORD environment variable not set")
        return False
    
    # Get summaries
    call_summary = get_call_summary(date_str)
    fax_summary = get_fax_summary(date_str)
    
    if not call_summary and not fax_summary:
        print("❌ Error: Could not load any report data")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['From'] = sender_email
        msg['To'] = ', '.join(receiver_emails)
        msg['Subject'] = f"📊 RingCentral Complete Activity Report - {date_str}"
        
        # Create HTML body
        html_body = create_email_body(date_str, call_summary, fax_summary)
        msg.attach(MIMEText(html_body, 'html'))
        
        # Attach call productivity report
        call_report = f'exports/{date_str}-nightangle-calls.xlsx'
        if os.path.exists(call_report):
            with open(call_report, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={date_str}-nightangle-calls.xlsx')
                msg.attach(part)
                print(f"✅ Attached: {call_report}")
        else:
            print(f"⚠️  Warning: {call_report} not found")
        
        # Attach fax analysis report
        fax_report = f'exports/fax_analysis_{date_str}.xlsx'
        if os.path.exists(fax_report):
            with open(fax_report, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename=fax_analysis_{date_str}.xlsx')
                msg.attach(part)
                print(f"✅ Attached: {fax_report}")
        else:
            print(f"⚠️  Warning: {fax_report} not found")
        
        # Send email
        print("📤 Sending email...")
        server = smtplib.SMTP('smtp.office365.com', 587)
        server.starttls()
        server.login(sender_email, password)
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Email sent successfully to:")
        for email in receiver_emails:
            print(f"   • {email}")
        print(f"📊 Email includes:")
        print(f"   • Comprehensive HTML analysis")
        print(f"   • Call productivity report (Excel)")
        print(f"   • Fax analysis report (Excel)")
        return True
        
    except Exception as e:
        print(f"❌ Failed to send email: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("📊 COMPREHENSIVE REPORT EMAIL SENDER")
    print("=" * 60)
    
    # Check if date is provided via command line arguments or environment variable
    if len(sys.argv) >= 4:
        # Date provided as command line arguments
        date_str = sys.argv[3]
        print(f"📅 Using date from command line: {date_str}")
    elif os.getenv('REPORT_DATE_STR'):
        # Date provided via environment variables
        date_str = os.getenv('REPORT_DATE_STR')
        print(f"📅 Using date from environment: {date_str}")
    else:
        # Default: yesterday
        yesterday = datetime.now() - timedelta(days=1)
        date_str = yesterday.strftime("%Y-%m-%d")
        print(f"📅 Using default date (yesterday): {date_str}")
    
    print(f"📅 Preparing reports for: {date_str}")
    
    # Check if reports exist (using dynamic date)
    call_report = f'exports/{date_str}-nightangle-calls.xlsx'
    fax_report = f'exports/fax_analysis_{date_str}.xlsx'
    
    if not os.path.exists(call_report):
        print(f"❌ Call report not found: {call_report}")
        print("💡 Run: python3 improved_call_logs.py")
        return
    
    if not os.path.exists(fax_report):
        print(f"❌ Fax report not found: {fax_report}")
        print("💡 Run: python3 analyze_fax_senders.py")
        return
    
    print("✅ Both reports found")
    
    # Send email
    success = send_email_with_reports(date_str)
    
    if success:
        print("\n🎉 Complete reports sent successfully!")
    else:
        print("\n❌ Failed to send reports")

if __name__ == "__main__":
    main()
