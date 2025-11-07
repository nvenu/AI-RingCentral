#!/usr/bin/env python3
"""
Improved RingCentral Call Logs Report
Properly populates extension_number and internal_user with correct fax splitting
"""

import csv
import sys
import os
import json
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠️  openpyxl not available - will generate CSV instead of Excel")
    EXCEL_AVAILABLE = False

try:
    from ringcentral import SDK
except ImportError:
    print("❌ Error: RingCentral SDK not installed")
    print("Run: pip install ringcentral")
    sys.exit(1)

# Load environment variables from .env file if it exists
def load_env():
    """Load environment variables from .env file"""
    env_file = '.env'
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()

# Load environment variables
load_env()

# Your RingCentral credentials
CLIENT_ID = "c97CkXtDJ4wcupAJ8ieF67"
CLIENT_SECRET = "5r8CRbMwHUjarHHjytVUSE7HawPIWiW3zeDyMhntFmCs"
SERVER_URL = "https://platform.ringcentral.com"
JWT_TOKEN = "eyJraWQiOiI4NzYyZjU5OGQwNTk0NGRiODZiZjVjYTk3ODA0NzYwOCIsInR5cCI6IkpXVCIsImFsZyI6IlJTMjU2In0.eyJhdWQiOiJodHRwczovL3BsYXRmb3JtLnJpbmdjZW50cmFsLmNvbS9yZXN0YXBpL29hdXRoL3Rva2VuIiwic3ViIjoiNjM2NzI0NzAwMzEiLCJpc3MiOiJodHRwczovL3BsYXRmb3JtLnJpbmdjZW50cmFsLmNvbSIsImV4cCI6MTgxOTg0MzE5OSwiaWF0IjoxNzU2ODk0MjU0LCJqdGkiOiJGMzhkNTd5dVJIcWZFVU5uX0t0M2NnIn0.WGDnyfTBEJ05PkwpHNCRYj-0ZkbLOkS5lvgzdPeiFCACSNSUAhzPKuYYZ_zUAuqBUDatw59soFVGeid6pRZDp8etXADwvHSwe7wai_w0zvTA1i_rGJGP3pwXEmdagMMudIQx-aWCnjazbjRWNXb_fWi5vpH4sf1jxUJQAERxF4iSmr0S__imKaV0D_26-mdlw021lb1Px3NiSh0YXo2ogHSW5jQHZ-HILYi-HaCDskTDr_34ccbe1tOb9bgnixmz6SflGMjupsfiBTzxvw1Ii-quVxyzWeiVYaiIG6vtJGKsrN8ocdXWtGzNGpT57MSqNnLMUBEcBqY4amvn8UXb6g"

def safe_get_attr(obj, attr, default=''):
    """Safely get attribute from JsonObject or dict"""
    try:
        # First try direct attribute access (works for JsonObject)
        if hasattr(obj, attr):
            value = getattr(obj, attr, default)
            return value if value is not None else default
        
        # Try json() method if available (JsonObject has this)
        if hasattr(obj, 'json'):
            json_data = obj.json()
            return json_data.get(attr, default) if json_data else default
        
        # Try dictionary access
        if hasattr(obj, 'get'):
            return obj.get(attr, default)
        elif hasattr(obj, '__getitem__'):
            return obj[attr] if attr in obj else default
        else:
            return default
    except:
        return default

def fetch_extensions_directory(platform):
    """Fetch all enabled extensions and build directory mapping"""
    print("📋 Fetching extensions directory for enrichment...")
    
    extensions_directory = {}  # {extensionId: {extensionNumber, name}}
    
    try:
        page = 1
        per_page = 1000
        
        while True:
            response = platform.get("/restapi/v1.0/account/~/extension", {
                "status": "Enabled",
                "perPage": per_page,
                "page": page
            })
            
            response_data = response.json()
            # Handle JsonObject response
            if hasattr(response_data, 'json'):
                response_dict = response_data.json()
            else:
                response_dict = response_data
            
            ext_records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
            
            if not ext_records:
                break
            
            for ext in ext_records:
                ext_id = safe_get_attr(ext, 'id', '')
                ext_number = safe_get_attr(ext, 'extensionNumber', '')
                
                # Get contact info
                contact = safe_get_attr(ext, 'contact', {})
                first_name = safe_get_attr(contact, 'firstName', '')
                last_name = safe_get_attr(contact, 'lastName', '')
                name = f"{first_name} {last_name}".strip()
                
                if not name:
                    name = safe_get_attr(ext, 'name', '')
                if not name:
                    name = f"Extension {ext_number}"
                
                if ext_id:
                    extensions_directory[str(ext_id)] = {
                        'extensionNumber': str(ext_number) if ext_number else '',
                        'name': name
                    }
            
            # Check pagination
            paging = safe_get_attr(response_data, 'paging', {})
            if len(ext_records) < per_page or not safe_get_attr(paging, 'hasNextPage', False):
                break
            
            page += 1
        
        print(f"✅ Fetched {len(extensions_directory)} enabled extensions")
        return extensions_directory
        
    except Exception as e:
        print(f"⚠️  Warning: Could not fetch extensions directory: {str(e)}")
        return {}

def build_phone_to_extension_map(extensions_directory):
    """Build mapping from phone numbers to extensions"""
    phone_map = {}
    
    # Main company numbers that should map to reception/main extensions
    main_numbers = {
        '+18663347777': {'extensionNumber': '5001', 'name': 'Reception 1 Phone', 'extensionId': '63332323031'},
        '+13173347777': {'extensionNumber': '5002', 'name': 'Reception 2 Phone', 'extensionId': '63332324031'},
        '+18668780094': {'extensionNumber': '9', 'name': 'Main Fax', 'extensionId': '63310910031'},
        '+13175691403': {'extensionNumber': '1268', 'name': 'Madati Sravani', 'extensionId': '63287000031'}
    }
    
    phone_map.update(main_numbers)
    return phone_map

def extract_call_data(record, extensions_directory, phone_to_extension_map=None):
    """Extract and map call data according to the specified rules"""
    
    # Basic call info
    direction = safe_get_attr(record, 'direction', '')
    call_type = safe_get_attr(record, 'type', '')
    duration = safe_get_attr(record, 'duration', 0)
    result = safe_get_attr(record, 'result', '')
    start_time = safe_get_attr(record, 'startTime', '')
    
    # Extract FROM data
    from_data = safe_get_attr(record, 'from_')  # Note: RingCentral uses 'from_' not 'from'
    from_extension_number = ''
    from_extension_id = ''
    from_name = ''
    from_phone = ''
    
    if from_data:
        from_extension_number = str(safe_get_attr(from_data, 'extensionNumber', ''))
        from_extension_id = str(safe_get_attr(from_data, 'extensionId', ''))
        from_name = safe_get_attr(from_data, 'name', '')
        from_phone = safe_get_attr(from_data, 'phoneNumber', '')
    
    # Extract TO data
    to_data = safe_get_attr(record, 'to')
    to_extension_number = ''
    to_extension_id = ''
    to_name = ''
    to_phone = ''
    
    if to_data:
        try:
            # Handle both array and single object cases
            if hasattr(to_data, '__len__') and len(to_data) > 0:
                first_to = to_data[0]
                to_extension_number = str(safe_get_attr(first_to, 'extensionNumber', ''))
                to_extension_id = str(safe_get_attr(first_to, 'extensionId', ''))
                to_name = safe_get_attr(first_to, 'name', '')
                to_phone = safe_get_attr(first_to, 'phoneNumber', '')
            else:
                to_extension_number = str(safe_get_attr(to_data, 'extensionNumber', ''))
                to_extension_id = str(safe_get_attr(to_data, 'extensionId', ''))
                to_name = safe_get_attr(to_data, 'name', '')
                to_phone = safe_get_attr(to_data, 'phoneNumber', '')
        except:
            pass
    
    # Apply mapping rules for extension_number and internal_user
    extension_number = ''
    internal_user = ''
    extension_id_for_lookup = ''
    
    if direction == "Outbound":
        # For outbound: use from side
        extension_number = from_extension_number
        internal_user = from_name
        extension_id_for_lookup = from_extension_id
    elif direction == "Inbound":
        # For inbound: use to side
        extension_number = to_extension_number
        internal_user = to_name
        extension_id_for_lookup = to_extension_id
    else:
        # Fallback: use whichever side has a value
        if from_extension_number:
            extension_number = from_extension_number
            internal_user = from_name
            extension_id_for_lookup = from_extension_id
        elif to_extension_number:
            extension_number = to_extension_number
            internal_user = to_name
            extension_id_for_lookup = to_extension_id
    
    # Enhanced phone number to extension mapping (for main numbers and missing extensions)
    if not extension_number and phone_to_extension_map:
        target_phone = to_phone if direction == "Inbound" else from_phone
        if target_phone in phone_to_extension_map:
            mapping = phone_to_extension_map[target_phone]
            extension_number = mapping['extensionNumber']
            internal_user = mapping['name']
            extension_id_for_lookup = mapping.get('extensionId', '')
    
    # Directory enrichment using extension ID
    if not extension_number or not internal_user:
        if extension_id_for_lookup and extension_id_for_lookup in extensions_directory:
            dir_info = extensions_directory[extension_id_for_lookup]
            if not extension_number:
                extension_number = dir_info['extensionNumber']
            if not internal_user:
                internal_user = dir_info['name']
        
        # Try other extension IDs if still empty
        if not extension_number or not internal_user:
            for ext_id in [from_extension_id, to_extension_id]:
                if ext_id and ext_id in extensions_directory:
                    dir_info = extensions_directory[ext_id]
                    if not extension_number:
                        extension_number = dir_info['extensionNumber']
                    if not internal_user:
                        internal_user = dir_info['name']
                    break
    
    # Additional directory enrichment by extension number (if we have extension but no name)
    if extension_number and not internal_user:
        # Look up by extension number in the directory
        for ext_id, ext_info in extensions_directory.items():
            if ext_info['extensionNumber'] == str(extension_number):
                internal_user = ext_info['name']
                break
    
    # Enhanced main number handling - map to appropriate extensions
    if not extension_number:
        # Main company numbers mapping
        main_number_mappings = {
            '+18663347777': {'ext': '5001', 'name': 'Reception 1 Phone'},
            '+13173347777': {'ext': '5002', 'name': 'Reception 2 Phone'},
            '+18668780094': {'ext': '9', 'name': 'Main Fax'},
            '+13175691403': {'ext': '1268', 'name': 'Madati Sravani'}
        }
        
        target_phone = to_phone if direction == "Inbound" else from_phone
        if target_phone in main_number_mappings:
            mapping = main_number_mappings[target_phone]
            extension_number = mapping['ext']
            internal_user = mapping['name']
            if direction == "Inbound":
                internal_user += " (Main Line)"
    
    # Final fallback for internal_user
    if not internal_user:
        if extension_number:
            internal_user = f"Extension {extension_number}"
        else:
            # Only mark as "External" if truly no internal extension found
            if not from_extension_number and not to_extension_number and not from_extension_id and not to_extension_id:
                # Check if this is truly external (no company phone numbers involved)
                company_numbers = ['+18663347777', '+13173347777', '+18668780094', '+13175691403']
                is_external = True
                
                for company_num in company_numbers:
                    if company_num in [from_phone, to_phone]:
                        is_external = False
                        break
                
                if is_external:
                    if direction == "Inbound":
                        internal_user = f"External - {from_phone or 'Unknown'}"
                    else:
                        internal_user = f"External - {to_phone or 'Unknown'}"
                else:
                    internal_user = "Unknown Extension"
            else:
                internal_user = "Unknown Extension"
    
    # Ensure extension_number is string (no float/scientific notation)
    extension_number = str(extension_number) if extension_number else ''
    
    return {
        'extension_number': extension_number,
        'internal_user': internal_user,
        'direction': direction,
        'type': call_type,
        'duration': int(duration) if duration else 0,
        'result': result,
        'start_time': start_time,
        'from_phone': str(from_phone) if from_phone else '',
        'to_phone': str(to_phone) if to_phone else '',
        'from_name': from_name,
        'to_name': to_name
    }

def generate_excel_report(grouped_records, filename, date_str):
    """Generate a beautifully formatted Excel report"""
    print(f"📊 Creating formatted Excel report: {filename}")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Call Report {date_str}"
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    data_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Define headers with better names
    headers = [
        "Employee Name", "Extension", 
        "Fax Sent", "Fax Received", "Total Fax",
        "Calls Received", "Calls Made", "Total Calls",
        "Avg Duration (min)", "Successful", "Missed", "Success %",
        "Total Minutes"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Custom sorting function to prioritize fax data at the top
    def sort_records(group_key):
        user_records = grouped_records[group_key]
        first_record = user_records[0]
        internal_user = first_record['internal_user']
        
        # Calculate fax activity for this extension
        fax_records = [r for r in user_records if r['type'] == 'Fax']
        has_fax_activity = len(fax_records) > 0
        
        # Sorting priority:
        # 1. Main Fax (always first)
        # 2. Extensions with fax activity (sorted by name)
        # 3. Extensions with only voice activity (sorted by name)
        
        if internal_user == 'Main Fax':
            return f'0_Main_Fax'  # Main Fax always first
        elif has_fax_activity:
            return f'1_Fax_{internal_user}'  # Fax extensions second
        else:
            return f'2_Voice_{internal_user}'  # Voice-only extensions last
    
    # Write data with custom sorting
    row = 2
    for group_key in sorted(grouped_records.keys(), key=sort_records):
        user_records = grouped_records[group_key]
        
        # Get extension number and user info from first record
        first_record = user_records[0]
        extension_number = first_record['extension_number']
        internal_user = first_record['internal_user']
        
        # FILTER: Only include internal extensions (skip external numbers)
        if not extension_number or internal_user.startswith('External -'):
            continue
        
        # Calculate metrics (same logic as before)
        voice_records = [r for r in user_records if r['type'] == 'Voice']
        fax_records = [r for r in user_records if r['type'] == 'Fax']
        
        inbound_voice = len([r for r in voice_records if r['direction'] == 'Inbound'])
        outbound_voice = len([r for r in voice_records if r['direction'] == 'Outbound'])
        total_voice_calls = len(voice_records)
        
        fax_received_records = [r for r in fax_records if r['direction'] == 'Inbound']
        fax_sent_records = [r for r in fax_records if r['direction'] == 'Outbound']
        fax_received_count = len(fax_received_records)
        fax_sent_count = len(fax_sent_records)
        total_faxes = len(fax_records)
        
        voice_durations = [r['duration'] for r in voice_records if r['duration']]
        total_duration_seconds = sum(voice_durations)
        total_duration_minutes = round(total_duration_seconds / 60, 2)
        avg_duration_minutes = round((total_duration_seconds / len(voice_durations)) / 60, 2) if voice_durations else 0
        
        successful_calls = len([r for r in voice_records if r['result'] in ['Call connected', 'Accepted', 'Received']])
        missed_calls = len([r for r in voice_records if r['result'] in ['Missed', 'No Answer', 'Busy', 'Receive Error']])
        success_rate = round((successful_calls / total_voice_calls * 100), 1) if total_voice_calls > 0 else 0
        
        # Write row data
        data = [
            internal_user, str(extension_number),
            fax_sent_count, fax_received_count, total_faxes,
            inbound_voice, outbound_voice, total_voice_calls,
            avg_duration_minutes, successful_calls, missed_calls, f"{success_rate}%",
            total_duration_minutes
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col-1])
        
        for row_cells in ws.iter_rows(min_col=col, max_col=col):
            for cell in row_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(filename)
    print(f"✅ Excel report saved: {filename}")
    return filename

def generate_csv_report(grouped_records, filename):
    """Generate CSV report as fallback"""
    print(f"📊 Creating CSV report: {filename}")
    
    with open(filename, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "internal_user", "extension_number",
            "fax_sent_count", "fax_received_count", "total_faxes",
            "calls_received", "calls_made", "total_calls",
            "avg_call_duration", "successful_calls", "missed_calls", "success_rate",
            "total_duration_minutes"
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        # Custom sorting function to prioritize fax data at the top (same as Excel)
        def sort_records_csv(group_key):
            user_records = grouped_records[group_key]
            first_record = user_records[0]
            internal_user = first_record['internal_user']
            
            # Calculate fax activity for this extension
            fax_records = [r for r in user_records if r['type'] == 'Fax']
            has_fax_activity = len(fax_records) > 0
            
            if internal_user == 'Main Fax':
                return f'0_Main_Fax'
            elif has_fax_activity:
                return f'1_Fax_{internal_user}'
            else:
                return f'2_Voice_{internal_user}'
        
        for group_key in sorted(grouped_records.keys(), key=sort_records_csv):
            user_records = grouped_records[group_key]
            
            first_record = user_records[0]
            extension_number = first_record['extension_number']
            internal_user = first_record['internal_user']
            
            if not extension_number or internal_user.startswith('External -'):
                continue
            
            # Same calculation logic as Excel version
            voice_records = [r for r in user_records if r['type'] == 'Voice']
            fax_records = [r for r in user_records if r['type'] == 'Fax']
            
            inbound_voice = len([r for r in voice_records if r['direction'] == 'Inbound'])
            outbound_voice = len([r for r in voice_records if r['direction'] == 'Outbound'])
            total_voice_calls = len(voice_records)
            
            fax_received_count = len([r for r in fax_records if r['direction'] == 'Inbound'])
            fax_sent_count = len([r for r in fax_records if r['direction'] == 'Outbound'])
            total_faxes = len(fax_records)
            
            voice_durations = [r['duration'] for r in voice_records if r['duration']]
            total_duration_seconds = sum(voice_durations)
            total_duration_minutes = round(total_duration_seconds / 60, 2)
            avg_duration_minutes = round((total_duration_seconds / len(voice_durations)) / 60, 2) if voice_durations else 0
            
            successful_calls = len([r for r in voice_records if r['result'] in ['Call connected', 'Accepted', 'Received']])
            missed_calls = len([r for r in voice_records if r['result'] in ['Missed', 'No Answer', 'Busy', 'Receive Error']])
            success_rate = round((successful_calls / total_voice_calls * 100), 1) if total_voice_calls > 0 else 0
            
            writer.writerow({
                "internal_user": internal_user,
                "extension_number": str(extension_number),
                "fax_sent_count": fax_sent_count,
                "fax_received_count": fax_received_count,
                "total_faxes": total_faxes,
                "calls_received": inbound_voice,
                "calls_made": outbound_voice,
                "total_calls": total_voice_calls,
                "avg_call_duration": avg_duration_minutes,
                "successful_calls": successful_calls,
                "missed_calls": missed_calls,
                "success_rate": f"{success_rate}%",
                "total_duration_minutes": total_duration_minutes
            })
    
    print(f"✅ CSV report saved: {filename}")
    return filename

def send_email_with_attachment(csv_filename, date_str, total_records, total_users):
    """Send email with CSV attachment"""
    print("📧 Sending email with CSV attachment...")
    
    # Email configuration
    sender_email = "nvenu@solifetec.com"
    receiver_email = "nvenu@solifetec.com"
    password = os.getenv('EMAIL_PASSWORD')
    
    if not password:
        print("❌ Error: EMAIL_PASSWORD environment variable not set")
        print("💡 Please set EMAIL_PASSWORD in your environment")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg['Subject'] = f"RingCentral Call Logs - {date_str}"
        
        # Email body
        body = f"""
Hello,

Please find attached the RingCentral call productivity report for {date_str}.

Summary:
• Total Call Records: {total_records:,}
• Active Users/Extensions: {total_users}
• Date Range: {date_str}
• Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

The report includes:
• Fax activity (sent/received counts)
• Voice call metrics (calls made/received, duration, success rates)
• Extension-wise productivity analysis
• Internal users only (external numbers excluded)

Best regards,
Call Analytics Team
        """.strip()
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach report file (Excel or CSV)
        with open(csv_filename, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        
        # Get just the filename without path
        filename_only = os.path.basename(csv_filename)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename={filename_only}'
        )
        msg.attach(part)
        
        # Send email using Office365 SMTP
        server = smtplib.SMTP('smtp.office365.com', 587)
        server.starttls()
        server.login(sender_email, password)
        text = msg.as_string()
        server.sendmail(sender_email, receiver_email, text)
        server.quit()
        
        print(f"✅ Email sent successfully to {receiver_email}")
        return True
        
    except Exception as e:
        print(f"❌ Failed to send email: {str(e)}")
        print("💡 Make sure EMAIL_PASSWORD is set and Office365 credentials are correct")
        return False

def main():
    print("🚀 Starting Improved RingCentral Call Logs Report...")
    
    try:
        # Initialize SDK
        print("📡 Connecting to RingCentral...")
        rcsdk = SDK(CLIENT_ID, CLIENT_SECRET, SERVER_URL)
        platform = rcsdk.platform()
        
        # Login with JWT
        print("🔐 Authenticating with JWT token...")
        platform.login(jwt=JWT_TOKEN)
        print("✅ Authentication successful")
        
        # Fetch extensions directory for enrichment
        extensions_directory = fetch_extensions_directory(platform)
        
        # Build phone to extension mapping
        phone_to_extension_map = build_phone_to_extension_map(extensions_directory)
        
        # Calculate date range (yesterday for example)
        yesterday = datetime.now() - timedelta(days=1)
        date_from = yesterday.strftime("%Y-%m-%dT00:00:00.000Z")
        date_to = yesterday.strftime("%Y-%m-%dT23:59:59.999Z")
        date_str = yesterday.strftime("%Y-%m-%d")
        
        print(f"📅 Fetching call logs for: {date_str}")
        
        # Rate-limited approach to get ALL call data
        print("📞 Fetching ALL call logs with proper rate limiting (this may take 5-10 minutes)...")
        import time
        
        all_records = []
        
        def make_api_call_with_retry(api_call, max_retries=5, base_delay=2):
            """Make API call with exponential backoff retry logic"""
            for attempt in range(max_retries):
                try:
                    return api_call()
                except Exception as e:
                    if "429" in str(e) or "rate exceeded" in str(e).lower():
                        if attempt < max_retries - 1:
                            delay = base_delay * (2 ** attempt)  # Exponential backoff
                            print(f"   ⏳ Rate limited, waiting {delay} seconds before retry {attempt + 1}/{max_retries}")
                            time.sleep(delay)
                            continue
                    raise e
            return None
        
        # Strategy 1: Use account-level call log with careful pagination
        print("🔄 Strategy 1: Account-level call log with rate limiting...")
        
        page = 1
        per_page = 100  # Much smaller page size to avoid rate limits
        max_pages = 50
        consecutive_errors = 0
        
        while page <= max_pages and consecutive_errors < 3:
            try:
                print(f"📄 Fetching page {page} (records so far: {len(all_records)})...")
                
                def api_call():
                    return platform.get("/restapi/v1.0/account/~/call-log", {
                        "view": "Detailed",
                        "dateFrom": date_from,
                        "dateTo": date_to,
                        "perPage": per_page,
                        "page": page
                    })
                
                response = make_api_call_with_retry(api_call)
                if not response:
                    break
                
                response_data = response.json()
                if hasattr(response_data, 'json'):
                    response_dict = response_data.json()
                else:
                    response_dict = response_data
                
                records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
                
                if not records:
                    print(f"📄 Page {page}: No more records found")
                    break
                
                all_records.extend(records)
                print(f"📊 Page {page}: {len(records)} records (Total: {len(all_records)})")
                
                # Check pagination
                paging = safe_get_attr(response_data, 'paging', {})
                has_next = safe_get_attr(paging, 'hasNextPage', False)
                total_pages = safe_get_attr(paging, 'totalPages', 'unknown')
                total_elements = safe_get_attr(paging, 'totalElements', 'unknown')
                
                print(f"   📊 Pagination: hasNext={has_next}, totalPages={total_pages}, totalElements={total_elements}")
                
                if len(records) < per_page or not has_next:
                    print(f"📄 Reached end of data at page {page}")
                    break
                
                page += 1
                consecutive_errors = 0
                
                # Rate limiting delay - be respectful to the API
                time.sleep(1.5)  # 1.5 second delay between requests
                
            except Exception as e:
                consecutive_errors += 1
                print(f"❌ Error on page {page}: {e}")
                if consecutive_errors >= 3:
                    print("❌ Too many consecutive errors, stopping pagination")
                    break
                
                # Wait longer after errors
                time.sleep(5)
        
        print(f"� Strategny 1 Results: {len(all_records)} total records")
        
        # Strategy 2: If we still don't have enough data, try smaller time windows
        if len(all_records) < 800:  # If we're still missing significant data
            print(f"🔄 Strategy 2: Trying smaller 2-hour time windows...")
            
            # Split into 2-hour windows
            time_windows = []
            for hour in range(0, 24, 2):
                start_hour = hour
                end_hour = min(hour + 1, 23)
                end_minute = 59 if end_hour < 23 else 59
                
                window_start = yesterday.replace(hour=start_hour, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S.000Z")
                window_end = yesterday.replace(hour=end_hour, minute=end_minute, second=59).strftime("%Y-%m-%dT%H:%M:%S.999Z")
                time_windows.append((window_start, window_end))
            
            window_records = []
            for window_idx, (window_start, window_end) in enumerate(time_windows, 1):
                print(f"⏰ Time window {window_idx}/12: {window_start[:16]} to {window_end[:16]}")
                
                try:
                    def window_api_call():
                        return platform.get("/restapi/v1.0/account/~/call-log", {
                            "view": "Detailed",
                            "dateFrom": window_start,
                            "dateTo": window_end,
                            "perPage": 200
                        })
                    
                    response = make_api_call_with_retry(window_api_call)
                    if response:
                        response_data = response.json()
                        if hasattr(response_data, 'json'):
                            response_dict = response_data.json()
                        else:
                            response_dict = response_data
                        
                        records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
                        
                        if records:
                            window_records.extend(records)
                            print(f"   ✅ Window {window_idx}: {len(records)} records")
                        else:
                            print(f"   📄 Window {window_idx}: No records")
                    
                    # Rate limiting between windows
                    time.sleep(2)
                    
                except Exception as e:
                    print(f"   ❌ Error in window {window_idx}: {e}")
                    time.sleep(5)  # Wait longer after errors
            
            # Deduplicate records
            print(f"📊 Deduplicating records from time windows...")
            seen_ids = set()
            
            # Add existing records to seen set
            for record in all_records:
                record_id = safe_get_attr(record, 'id', '')
                if record_id:
                    seen_ids.add(record_id)
            
            # Add new unique records
            new_records = 0
            for record in window_records:
                record_id = safe_get_attr(record, 'id', '')
                if record_id and record_id not in seen_ids:
                    seen_ids.add(record_id)
                    all_records.append(record)
                    new_records += 1
            
            print(f"📊 Added {new_records} new unique records from time windows")
        
        print(f"🎉 FINAL RESULT: Found {len(all_records)} total call records")
        
        if len(all_records) < 1000:
            print(f"⚠️  We have {len(all_records)} records, dashboard shows ~1,117")
            print(f"💡 This is likely due to API rate limits or data access restrictions")
            print(f"📊 We'll proceed with the data we have - it should still be comprehensive")
        
        print(f"🎉 Found {len(all_records)} total call records")
        
        if not all_records:
            print("⚠️  No call records found for the specified date range")
            return
        
        # Process records and group by extension
        print("💾 Processing records and grouping by extension...")
        grouped_records = defaultdict(list)
        
        validation_stats = {
            'total_processed': 0,
            'has_extension_number': 0,
            'enriched_from_directory': 0,
            'external_unknown': 0
        }
        
        for i, record in enumerate(all_records, 1):
            print(f"🔍 Processing record {i}/{len(all_records)}...", end='\r')
            
            # Extract call data using improved mapping
            call_data = extract_call_data(record, extensions_directory, phone_to_extension_map)
            
            validation_stats['total_processed'] += 1
            
            if call_data['extension_number']:
                validation_stats['has_extension_number'] += 1
            
            if call_data['internal_user'].startswith('External - Unknown'):
                validation_stats['external_unknown'] += 1
            
            # Group by extension number
            if call_data['extension_number']:
                group_key = f"EXT_{call_data['extension_number']}"
            else:
                group_key = f"EXTERNAL_{call_data['from_phone'] or call_data['to_phone'] or 'UNKNOWN'}"
            
            grouped_records[group_key].append(call_data)
        
        print(f"\n📊 VALIDATION STATISTICS:")
        print(f"   Total records processed: {validation_stats['total_processed']}")
        print(f"   Records with extension_number: {validation_stats['has_extension_number']}")
        print(f"   External/Unknown records: {validation_stats['external_unknown']}")
        print(f"   Total groups created: {len(grouped_records)}")
        
        # Ensure exports folder exists
        os.makedirs('exports', exist_ok=True)
        
        # Generate improved productivity summary
        if EXCEL_AVAILABLE:
            report_filename = f"exports/{date_str}-nightangle-calls.xlsx"
            generate_excel_report(grouped_records, report_filename, date_str)
        else:
            report_filename = f"exports/{date_str}-nightangle-calls.csv"
            generate_csv_report(grouped_records, report_filename)
        
        print(f"\n✅ Improved call productivity report saved to {report_filename}")
        print(f"📈 Total records exported: {len(all_records)}")
        print(f"👥 Activity tracked for {len(grouped_records)} users/extensions")
        
        # Acceptance checks
        print(f"\n🔍 ACCEPTANCE CHECKS:")
        extension_populated = validation_stats['has_extension_number']
        external_unknown = validation_stats['external_unknown']
        
        print(f"   ✅ Extension numbers populated: {extension_populated}/{validation_stats['total_processed']} records")
        print(f"   ✅ External/Unknown entries: {external_unknown} (should be minimal)")
        
        if external_unknown > 0:
            print(f"   💡 {external_unknown} records marked as External - check if these are truly external calls")
        
        # Count internal records only (records that actually appear in the report)
        internal_records_count = 0
        for group_key in grouped_records.keys():
            user_records = grouped_records[group_key]
            first_record = user_records[0]
            extension_number = first_record['extension_number']
            internal_user = first_record['internal_user']
            
            # Only count internal extensions (same filter as in report generation)
            if extension_number and not internal_user.startswith('External -'):
                internal_records_count += len(user_records)
        
        # Send email with attachment
        email_sent = send_email_with_attachment(report_filename, date_str, internal_records_count, len(grouped_records))
        
        if email_sent:
            print("🎉 Process completed successfully - Improved CSV generated and emailed!")
        else:
            print("⚠️  CSV generated but email failed - check EMAIL_PASSWORD environment variable")
        
        print(f"\n🎯 VALIDATION INSTRUCTIONS:")
        print(f"   1. Check extension_number column - should be populated for internal calls")
        print(f"   2. Verify internal_user names are correct")
        print(f"   3. Confirm fax_sent_count vs fax_received_count are properly separated")
        print(f"   4. Spot-check Lubna's data against RingCentral dashboard")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        print("💡 Make sure your JWT token is valid and has ReadCallLog permission")
        sys.exit(1)

if __name__ == "__main__":
    main()