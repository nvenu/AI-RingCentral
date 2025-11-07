#!/usr/bin/env python3
"""
Analyze Fax Senders - Show who sent/received faxes through main fax line
"""

import sys
import os
from datetime import datetime, timedelta
from collections import defaultdict

try:
    from ringcentral import SDK
except ImportError:
    print("❌ Error: RingCentral SDK not installed")
    print("Run: pip install ringcentral")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("⚠️  openpyxl not available - will generate CSV instead")
    EXCEL_AVAILABLE = False

# Load environment variables
def load_env():
    env_file = '.env'
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()

load_env()

# RingCentral credentials
CLIENT_ID = "c97CkXtDJ4wcupAJ8ieF67"
CLIENT_SECRET = "5r8CRbMwHUjarHHjytVUSE7HawPIWiW3zeDyMhntFmCs"
SERVER_URL = "https://platform.ringcentral.com"
JWT_TOKEN = "eyJraWQiOiI4NzYyZjU5OGQwNTk0NGRiODZiZjVjYTk3ODA0NzYwOCIsInR5cCI6IkpXVCIsImFsZyI6IlJTMjU2In0.eyJhdWQiOiJodHRwczovL3BsYXRmb3JtLnJpbmdjZW50cmFsLmNvbS9yZXN0YXBpL29hdXRoL3Rva2VuIiwic3ViIjoiNjM2NzI0NzAwMzEiLCJpc3MiOiJodHRwczovL3BsYXRmb3JtLnJpbmdjZW50cmFsLmNvbSIsImV4cCI6MTgxOTg0MzE5OSwiaWF0IjoxNzU2ODk0MjU0LCJqdGkiOiJGMzhkNTd5dVJIcWZFVU5uX0t0M2NnIn0.WGDnyfTBEJ05PkwpHNCRYj-0ZkbLOkS5lvgzdPeiFCACSNSUAhzPKuYYZ_zUAuqBUDatw59soFVGeid6pRZDp8etXADwvHSwe7wai_w0zvTA1i_rGJGP3pwXEmdagMMudIQx-aWCnjazbjRWNXb_fWi5vpH4sf1jxUJQAERxF4iSmr0S__imKaV0D_26-mdlw021lb1Px3NiSh0YXo2ogHSW5jQHZ-HILYi-HaCDskTDr_34ccbe1tOb9bgnixmz6SflGMjupsfiBTzxvw1Ii-quVxyzWeiVYaiIG6vtJGKsrN8ocdXWtGzNGpT57MSqNnLMUBEcBqY4amvn8UXb6g"

def safe_get_attr(obj, attr, default=''):
    """Safely get attribute from JsonObject or dict"""
    try:
        if hasattr(obj, attr):
            value = getattr(obj, attr, default)
            return value if value is not None else default
        if hasattr(obj, 'json'):
            json_data = obj.json()
            return json_data.get(attr, default) if json_data else default
        if hasattr(obj, 'get'):
            return obj.get(attr, default)
        elif hasattr(obj, '__getitem__'):
            return obj[attr] if attr in obj else default
        else:
            return default
    except:
        return default

def fetch_extensions_directory(platform):
    """Fetch all enabled extensions"""
    print("📋 Fetching extensions directory...")
    extensions_directory = {}
    
    try:
        page = 1
        per_page = 1000
        
        while True:
            response = platform.get("/restapi/v1.0/account/~/extension", {
                "status": "Enabled",
                "perPage": per_page,
                "page": page
            })
            
            response_data = response.json()
            if hasattr(response_data, 'json'):
                response_dict = response_data.json()
            else:
                response_dict = response_data
            
            ext_records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
            
            if not ext_records:
                break
            
            for ext in ext_records:
                ext_id = safe_get_attr(ext, 'id', '')
                ext_number = safe_get_attr(ext, 'extensionNumber', '')
                
                contact = safe_get_attr(ext, 'contact', {})
                first_name = safe_get_attr(contact, 'firstName', '')
                last_name = safe_get_attr(contact, 'lastName', '')
                name = f"{first_name} {last_name}".strip()
                
                if not name:
                    name = safe_get_attr(ext, 'name', '')
                if not name:
                    name = f"Extension {ext_number}"
                
                if ext_id:
                    extensions_directory[str(ext_id)] = {
                        'extensionNumber': str(ext_number) if ext_number else '',
                        'name': name
                    }
            
            paging = safe_get_attr(response_data, 'paging', {})
            if len(ext_records) < per_page or not safe_get_attr(paging, 'hasNextPage', False):
                break
            
            page += 1
        
        print(f"✅ Fetched {len(extensions_directory)} enabled extensions")
        return extensions_directory
        
    except Exception as e:
        print(f"⚠️  Warning: Could not fetch extensions directory: {str(e)}")
        return {}

def extract_fax_data(record, extensions_directory):
    """Extract fax data from a call record"""
    try:
        direction = safe_get_attr(record, 'direction', '')
        start_time = safe_get_attr(record, 'startTime', '')
        result = safe_get_attr(record, 'result', '')
        
        # FROM data
        from_data = safe_get_attr(record, 'from_')
        from_ext_number = str(safe_get_attr(from_data, 'extensionNumber', ''))
        from_ext_id = str(safe_get_attr(from_data, 'extensionId', ''))
        from_name = safe_get_attr(from_data, 'name', '')
        from_phone = safe_get_attr(from_data, 'phoneNumber', '')
        
        # TO data
        to_data = safe_get_attr(record, 'to')
        to_ext_number = ''
        to_ext_id = ''
        to_name = ''
        to_phone = ''
        
        if to_data:
            try:
                if hasattr(to_data, '__len__') and len(to_data) > 0:
                    first_to = to_data[0]
                    to_ext_number = str(safe_get_attr(first_to, 'extensionNumber', ''))
                    to_ext_id = str(safe_get_attr(first_to, 'extensionId', ''))
                    to_name = safe_get_attr(first_to, 'name', '')
                    to_phone = safe_get_attr(first_to, 'phoneNumber', '')
                else:
                    to_ext_number = str(safe_get_attr(to_data, 'extensionNumber', ''))
                    to_ext_id = str(safe_get_attr(to_data, 'extensionId', ''))
                    to_name = safe_get_attr(to_data, 'name', '')
                    to_phone = safe_get_attr(to_data, 'phoneNumber', '')
            except:
                pass
        
        # Enrich with directory data
        if from_ext_id and from_ext_id in extensions_directory:
            dir_info = extensions_directory[from_ext_id]
            if not from_name:
                from_name = dir_info['name']
            if not from_ext_number:
                from_ext_number = dir_info['extensionNumber']
        
        if to_ext_id and to_ext_id in extensions_directory:
            dir_info = extensions_directory[to_ext_id]
            if not to_name:
                to_name = dir_info['name']
            if not to_ext_number:
                to_ext_number = dir_info['extensionNumber']
        
        # Determine who actually sent/received the fax
        if direction == "Outbound":
            sender_name = from_name or f"Ext {from_ext_number}" or "Unknown"
            sender_ext = from_ext_number
            recipient = to_phone or to_name or "Unknown"
        else:  # Inbound
            sender_name = from_phone or from_name or "Unknown"
            sender_ext = ""
            # Better recipient identification for inbound faxes
            if to_ext_number == '9':
                recipient = "Main Fax (Ext 9)"
            elif to_name:
                recipient = to_name
            elif to_ext_number:
                recipient = f"Ext {to_ext_number}"
            else:
                recipient = to_phone or "Unknown"
        
        return {
            'timestamp': start_time,
            'direction': direction,
            'sender_name': sender_name,
            'sender_extension': sender_ext,
            'recipient': recipient,
            'from_phone': from_phone,
            'to_phone': to_phone,
            'from_ext': from_ext_number,
            'to_ext': to_ext_number,
            'result': result
        }
    except Exception as e:
        print(f"⚠️  Error extracting fax data: {e}")
        return None

def analyze_fax_details(platform, extensions_directory, date_from, date_to):
    """Analyze fax records in detail to show who sent/received them"""
    print(f"📠 Analyzing fax records in detail...")
    
    import time
    
    all_fax_records = []
    
    # Strategy 1: Try to get all fax records with pagination
    print("🔄 Strategy 1: Fetching fax records with pagination...")
    page = 1
    per_page = 100
    consecutive_errors = 0
    
    while page <= 50 and consecutive_errors < 3:
        try:
            print(f"📄 Fetching page {page} (faxes so far: {len(all_fax_records)})...")
            
            response = platform.get("/restapi/v1.0/account/~/call-log", {
                "view": "Detailed",
                "dateFrom": date_from,
                "dateTo": date_to,
                "type": "Fax",  # Only get fax records
                "perPage": per_page,
                "page": page
            })
            
            response_data = response.json()
            if hasattr(response_data, 'json'):
                response_dict = response_data.json()
            else:
                response_dict = response_data
            
            records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
            
            if not records:
                print(f"📄 Page {page}: No more records")
                break
            
            print(f"   Found {len(records)} fax records on page {page}")
            
            # Check pagination info
            paging = safe_get_attr(response_data, 'paging', {})
            has_next = safe_get_attr(paging, 'hasNextPage', False)
            
            if len(records) < per_page or not has_next:
                # Process these records before breaking
                for record in records:
                    fax_data = extract_fax_data(record, extensions_directory)
                    if fax_data:
                        all_fax_records.append(fax_data)
                print(f"📄 Reached end of data at page {page}")
                break
            # Process records
            for record in records:
                fax_data = extract_fax_data(record, extensions_directory)
                if fax_data:
                    all_fax_records.append(fax_data)
            
            page += 1
            consecutive_errors = 0
            time.sleep(1.5)
            
        except Exception as e:
            consecutive_errors += 1
            print(f"❌ Error on page {page}: {e}")
            if consecutive_errors >= 3:
                print("❌ Too many consecutive errors, stopping")
                break
            time.sleep(5)
    
    print(f"📊 Strategy 1 Results: {len(all_fax_records)} fax records")
    
    # Strategy 2: If we didn't get enough, try time windows
    if len(all_fax_records) < 400:
        print(f"\n🔄 Strategy 2: Trying 2-hour time windows to get all faxes...")
        
        from datetime import datetime, timedelta
        date_obj = datetime.strptime(date_from[:10], "%Y-%m-%d")
        
        time_windows = []
        for hour in range(0, 24, 2):
            start_hour = hour
            end_hour = min(hour + 1, 23)
            end_minute = 59 if end_hour < 23 else 59
            
            window_start = date_obj.replace(hour=start_hour, minute=0, second=0).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            window_end = date_obj.replace(hour=end_hour, minute=end_minute, second=59).strftime("%Y-%m-%dT%H:%M:%S.999Z")
            time_windows.append((window_start, window_end))
        
        window_records = []
        for window_idx, (window_start, window_end) in enumerate(time_windows, 1):
            print(f"⏰ Time window {window_idx}/12: {window_start[:16]} to {window_end[:16]}")
            
            try:
                response = platform.get("/restapi/v1.0/account/~/call-log", {
                    "view": "Detailed",
                    "dateFrom": window_start,
                    "dateTo": window_end,
                    "type": "Fax",
                    "perPage": 200
                })
                
                response_data = response.json()
                if hasattr(response_data, 'json'):
                    response_dict = response_data.json()
                else:
                    response_dict = response_data
                
                records = response_dict.get('records', []) if isinstance(response_dict, dict) else safe_get_attr(response_data, 'records', [])
                
                if records:
                    for record in records:
                        fax_data = extract_fax_data(record, extensions_directory)
                        if fax_data:
                            window_records.append(fax_data)
                    print(f"   ✅ Window {window_idx}: {len(records)} fax records")
                else:
                    print(f"   📄 Window {window_idx}: No records")
                
                time.sleep(2)
                
            except Exception as e:
                print(f"   ❌ Error in window {window_idx}: {e}")
                time.sleep(5)
        
        # Deduplicate
        print(f"📊 Deduplicating fax records...")
        seen_ids = set()
        unique_records = []
        
        for record in all_fax_records + window_records:
            record_id = f"{record['timestamp']}_{record['direction']}_{record['sender_name']}"
            if record_id not in seen_ids:
                seen_ids.add(record_id)
                unique_records.append(record)
        
        all_fax_records = unique_records
        print(f"📊 Total unique fax records: {len(all_fax_records)}")
    
    return all_fax_records

def generate_fax_report(fax_records, filename, date_str):
    """Generate detailed fax report"""
    print(f"📊 Creating fax analysis report...")
    
    if EXCEL_AVAILABLE:
        wb = Workbook()
        
        # Sheet 1: Fax by Sender
        ws1 = wb.active
        ws1.title = "Faxes by Sender"
        
        # Group by sender
        sender_stats = defaultdict(lambda: {'sent': 0, 'received': 0, 'extension': ''})
        
        for record in fax_records:
            if record['direction'] == 'Outbound':
                sender = record['sender_name']
                sender_stats[sender]['sent'] += 1
                sender_stats[sender]['extension'] = record['sender_extension']
            else:
                recipient = record['recipient']
                sender_stats[recipient]['received'] += 1
        
        # Write headers
        headers1 = ['Employee Name', 'Extension', 'Faxes Sent', 'Faxes Received', 'Total Faxes']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Write data
        row = 2
        for sender in sorted(sender_stats.keys()):
            stats = sender_stats[sender]
            ws1.cell(row=row, column=1, value=sender)
            ws1.cell(row=row, column=2, value=stats['extension'])
            ws1.cell(row=row, column=3, value=stats['sent'])
            ws1.cell(row=row, column=4, value=stats['received'])
            ws1.cell(row=row, column=5, value=stats['sent'] + stats['received'])
            row += 1
        
        # Sheet 2: Detailed Fax Log
        ws2 = wb.create_sheet("Detailed Fax Log")
        
        headers2 = ['Date/Time', 'Direction', 'Sender', 'Sender Ext', 'Recipient', 'From Phone', 'To Phone', 'Status']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Write detailed records
        row = 2
        for record in sorted(fax_records, key=lambda x: x['timestamp'], reverse=True):
            ws2.cell(row=row, column=1, value=record['timestamp'])
            ws2.cell(row=row, column=2, value=record['direction'])
            ws2.cell(row=row, column=3, value=record['sender_name'])
            ws2.cell(row=row, column=4, value=record['sender_extension'])
            ws2.cell(row=row, column=5, value=record['recipient'])
            ws2.cell(row=row, column=6, value=record['from_phone'])
            ws2.cell(row=row, column=7, value=record['to_phone'])
            ws2.cell(row=row, column=8, value=record['result'])
            row += 1
        
        # Auto-adjust column widths for both sheets
        for ws in [ws1, ws2]:
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                for cell in ws[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        wb.save(filename)
        print(f"✅ Excel report saved: {filename}")
    else:
        # CSV fallback
        import csv
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Date/Time', 'Direction', 'Sender', 'Sender Ext', 'Recipient', 'From Phone', 'To Phone', 'Status'])
            for record in sorted(fax_records, key=lambda x: x['timestamp'], reverse=True):
                writer.writerow([
                    record['timestamp'],
                    record['direction'],
                    record['sender_name'],
                    record['sender_extension'],
                    record['recipient'],
                    record['from_phone'],
                    record['to_phone'],
                    record['result']
                ])
        print(f"✅ CSV report saved: {filename}")

def main():
    print("📠 FAX SENDER ANALYSIS")
    print("=" * 60)
    
    try:
        # Initialize SDK
        print("📡 Connecting to RingCentral...")
        rcsdk = SDK(CLIENT_ID, CLIENT_SECRET, SERVER_URL)
        platform = rcsdk.platform()
        
        print("🔐 Authenticating...")
        platform.login(jwt=JWT_TOKEN)
        print("✅ Authentication successful")
        
        # Fetch extensions directory
        extensions_directory = fetch_extensions_directory(platform)
        
        # Date range (yesterday)
        yesterday = datetime.now() - timedelta(days=1)
        date_from = yesterday.strftime("%Y-%m-%dT00:00:00.000Z")
        date_to = yesterday.strftime("%Y-%m-%dT23:59:59.999Z")
        date_str = yesterday.strftime("%Y-%m-%d")
        
        print(f"📅 Analyzing faxes for: {date_str}")
        
        # Analyze fax records
        fax_records = analyze_fax_details(platform, extensions_directory, date_from, date_to)
        
        print(f"\n📊 SUMMARY:")
        print(f"   Total fax records: {len(fax_records)}")
        
        sent_count = len([r for r in fax_records if r['direction'] == 'Outbound'])
        received_count = len([r for r in fax_records if r['direction'] == 'Inbound'])
        
        print(f"   Faxes sent: {sent_count}")
        print(f"   Faxes received: {received_count}")
        
        # Generate report
        os.makedirs('exports', exist_ok=True)
        
        if EXCEL_AVAILABLE:
            filename = f"exports/fax_analysis_{date_str}.xlsx"
        else:
            filename = f"exports/fax_analysis_{date_str}.csv"
        
        generate_fax_report(fax_records, filename, date_str)
        
        print(f"\n✅ Fax analysis complete!")
        print(f"📁 Report saved: {filename}")
        print(f"\n💡 The report shows:")
        print(f"   • Who sent each fax (with their extension)")
        print(f"   • Who received each fax")
        print(f"   • Complete fax log with timestamps")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
